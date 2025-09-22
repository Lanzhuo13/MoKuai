import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import os

# ------------------- 路径配置 -------------------
class PathConfig:
    INPUT_FILE = "D:\\桌面\\表格生成\\备货单.xlsx"
    OUTPUT_DIR = "D:\\桌面\\表格生成"
    OUTPUT_PREFIX = "简洁备货单"  # 修改为简洁命名前缀
    USE_TIMESTAMP = True

    @classmethod
    def get_output_path(cls):
        if not os.path.exists(cls.OUTPUT_DIR):
            os.makedirs(cls.OUTPUT_DIR)
        # 时间戳格式修改为 月日时分 (MMDDHHMM)
        timestamp = datetime.now().strftime("%m%d%H%M") if cls.USE_TIMESTAMP else ""
        filename = f"{cls.OUTPUT_PREFIX}{timestamp}.xlsx" if timestamp else f"{cls.OUTPUT_PREFIX}.xlsx"
        return os.path.join(cls.OUTPUT_DIR, filename)

# ------------------- 数据标准化工具类 -------------------
class DataStandardizer:
    def __init__(self):
        self.field_mapping = {
            "类型": ["类型", "宝贝类型", "商品信息", "产品类型"],
            "图案": ["图案", "图案名称", "花纹", "印花"],
            "颜色": ["颜色", "色彩", "色系"],
            "规格": ["规格", "尺寸", "尺码"],
            "数量": ["数量", "件数", "总数", "库存"]
        }

    def standardize(self, file_path):
        try:
            df = pd.read_excel(file_path)
        except FileNotFoundError:
            raise FileNotFoundError(f"输入文件不存在：{file_path}")
        except Exception as e:
            raise RuntimeError(f"读取文件失败：{str(e)}")
        
        standardized_cols = {}
        for standard_col, aliases in self.field_mapping.items():
            matched_alias = None
            for alias in aliases:
                if alias.lower() in [col.lower() for col in df.columns]:
                    for col in df.columns:
                        if col.lower() == alias.lower():
                            matched_alias = col
                            break
                    break
            if not matched_alias:
                raise ValueError(
                    f"数据源缺少必要字段 '{standard_col}'，允许的别名包括：{', '.join(aliases)}"
                )
            standardized_cols[matched_alias] = standard_col
        
        df = df.rename(columns=standardized_cols)
        df["数量"] = pd.to_numeric(df["数量"], errors="coerce").fillna(0).astype(int)
        return df

# ------------------- 表格生成核心类 -------------------
class CompleteTableGenerator:
    def __init__(self):
        # 样式定义
        self.bold_font = Font(bold=True, size=11)
        self.normal_font = Font(size=11)
        self.title_total_fill = PatternFill(fgColor="538DD5", fill_type="solid")  # 标题色
        self.header_fill = PatternFill(fgColor="B8CCE4", fill_type="solid")       # 表头色
        self.white_fill = PatternFill(fgColor="FFFFFF", fill_type="solid")        # 白色
        self.center_align = Alignment(horizontal="center", vertical="center")
        
        # 边框样式（区块内完整框线，包括首行）
        side = Side(style="thin", color="000000")
        self.block_border = Border(
            left=side, right=side, top=side, bottom=side
        )
        # 无框线样式（区块外使用）
        self.no_border = Border(
            left=Side(style=None), right=Side(style=None),
            top=Side(style=None), bottom=Side(style=None)
        )
        
        # 常量定义
        self.min_columns = 6   # 最小列数
        self.max_columns = 10  # 最大列数
        self.page_max_rows = 50  # 每页最大行数
        self.left_start_col = 1    # 左侧表格首列
        self.right_start_col = 12  # 右侧表格首列
        self.space_col_width = 1   # 间隔列宽度
        self.block_spacing = 2     # 区块间垂直间隔行数

    def _get_formatted_date(self):
        return datetime.now().strftime("%Y年%m月%d日")

    def _sort_specs(self, specs):
        try:
            return sorted(specs, key=lambda x: float(x))
        except (ValueError, TypeError):
            return sorted(specs)

    # ------------------- 类型汇总表生成逻辑 -------------------
    def _calculate_type_aggregation(self, df):
        type_groups = df.groupby("类型", as_index=False)
        all_type_data = {}
        date_str = self._get_formatted_date()

        for _, group in type_groups:
            type_name = group["类型"].iloc[0]
            colors = sorted(group["颜色"].unique())
            raw_specs = sorted(group["规格"].unique())
            sorted_specs = self._sort_specs(raw_specs)

            headers = ["颜色"] + sorted_specs + ["行合计"]
            if len(headers) < self.min_columns:
                headers += [""] * (self.min_columns - len(headers))

            data_matrix = []
            for color in colors:
                color_data = group[group["颜色"] == color]
                row = [color]
                for spec in sorted_specs:
                    qty = color_data[color_data["规格"] == spec]["数量"].sum()
                    row.append(int(qty) if not pd.isna(qty) else 0)
                row.append(sum(row[1:len(sorted_specs)+1]))
                row += [""] * (len(headers) - len(row))
                data_matrix.append(row)

            all_type_data[type_name] = {
                "title": f"{type_name} - {date_str}",
                "headers": headers,
                "data": data_matrix,
                "specs": sorted_specs,
                "total_cols": len(headers),
                "data_rows": len(colors),
                "total_col_idx": len(sorted_specs) + 2
            }
        return all_type_data

    def _generate_original_type_sheet(self, ws, aggregation_data):
        ws.title = "类型数据汇总"
        current_row = 1

        sorted_types = sorted(
            aggregation_data.items(),
            key=lambda x: x[1]["total_cols"],
            reverse=True
        )

        for type_name, type_data in sorted_types:
            title_row = current_row  # 区块第一行（标题行）
            header_row = current_row + 1
            data_start = header_row + 1
            data_end = data_start + type_data["data_rows"] - 1
            total_row = data_end + 1
            total_cols = type_data["total_cols"]

            # 标题行（区块第一行，确保完整框线）
            ws.merge_cells(start_row=title_row, start_column=1,
                          end_row=title_row, end_column=total_cols)
            title_cell = ws.cell(row=title_row, column=1, value=type_data["title"])
            title_cell.font = self.bold_font
            title_cell.fill = self.title_total_fill
            title_cell.alignment = self.center_align
            title_cell.border = self.block_border  # 明确应用完整框线

            # 表头行
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=header_row, column=col, value=type_data["headers"][col-1])
                cell.font = self.bold_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align
                cell.border = self.block_border

            # 数据行
            for row_offset, data_row in enumerate(type_data["data"]):
                data_row_num = data_start + row_offset
                fill = self.header_fill if (row_offset % 2 == 1) else self.white_fill
                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=data_row_num, column=col, 
                                  value=data_row[col-1] if (col-1) < len(data_row) else "")
                    if col == 1 or col == type_data["total_col_idx"]:
                        cell.font = self.bold_font
                    else:
                        cell.font = self.normal_font
                    cell.fill = fill
                    cell.alignment = self.center_align
                    cell.border = self.block_border

            # 合计行
            ws.cell(row=total_row, column=1, value="合计").font = self.bold_font
            ws.cell(row=total_row, column=1).fill = self.title_total_fill
            ws.cell(row=total_row, column=1).alignment = self.center_align
            ws.cell(row=total_row, column=1).border = self.block_border

            # 规格列合计
            for col_offset in range(len(type_data["specs"])):
                col_idx = 2 + col_offset
                col_letter = get_column_letter(col_idx)
                ws.cell(row=total_row, column=col_idx,
                       value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})").font = self.bold_font
                ws.cell(row=total_row, column=col_idx).fill = self.title_total_fill
                ws.cell(row=total_row, column=col_idx).alignment = self.center_align
                ws.cell(row=total_row, column=col_idx).border = self.block_border

            # 总合计
            total_col_letter = get_column_letter(type_data["total_col_idx"])
            ws.cell(row=total_row, column=type_data["total_col_idx"],
                   value=f"=SUM({total_col_letter}{data_start}:{total_col_letter}{data_end})").font = self.bold_font
            ws.cell(row=total_row, column=type_data["total_col_idx"]).fill = self.title_total_fill
            ws.cell(row=total_row, column=type_data["total_col_idx"]).alignment = self.center_align
            ws.cell(row=total_row, column=type_data["total_col_idx"]).border = self.block_border

            # 空白填充列
            for col_idx in range(1, total_cols + 1):
                if (col_idx == 1 
                    or (2 <= col_idx <= 1 + len(type_data["specs"])) 
                    or col_idx == type_data["total_col_idx"]):
                    continue
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = self.title_total_fill
                cell.border = self.block_border
                cell.alignment = self.center_align

            # 样式设置
            for row in range(title_row, total_row + 1):
                ws.row_dimensions[row].height = 24
            for col in range(1, total_cols + 1):
                ws.column_dimensions[get_column_letter(col)].width = 7.5

            # 区块外间隔行（无框线）
            for row in range(total_row + 1, total_row + self.block_spacing):
                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.no_border
            current_row = total_row + self.block_spacing

    # ------------------- 图案明细工作表生成逻辑 -------------------
    def _get_pattern_table_structure(self, pattern_data):
        colors = sorted(pattern_data["颜色"].unique())
        sorted_specs = sorted(pattern_data["规格"].unique())

        base_headers = ["颜色"] + sorted_specs + ["行合计"]
        base_col_count = len(base_headers)
        
        if base_col_count < self.min_columns:
            total_cols = self.min_columns
            headers = base_headers + [""] * (self.min_columns - base_col_count)
        elif base_col_count > self.max_columns:
            total_cols = self.max_columns
            headers = base_headers[:self.max_columns]
            headers[-1] = "行合计"
        else:
            total_cols = base_col_count
            headers = base_headers

        total_col_idx = len(sorted_specs) + 2
        if total_col_idx > total_cols:
            total_col_idx = total_cols

        data_matrix = []
        for color in colors:
            color_data = pattern_data[pattern_data["颜色"] == color]
            row = [color]
            for spec in sorted_specs:
                qty = color_data[color_data["规格"] == spec]["数量"].sum()
                row.append(int(qty) if not pd.isna(qty) else 0)
            row_total = sum(row[1:len(sorted_specs)+1])
            row.append(row_total)
            row = row[:total_cols] + [""] * (total_cols - len(row))
            data_matrix.append(row)

        return {
            "headers": headers,
            "data": data_matrix,
            "row_count": len(colors),
            "total_cols": total_cols,
            "spec_count": len(sorted_specs),
            "total_col_idx": total_col_idx
        }

    def _sort_patterns_by_columns(self, patterns, order="asc"):
        patterns_with_cols = []
        for p in patterns:
            struct = self._get_pattern_table_structure(p["data"])
            patterns_with_cols.append({
                "name": p["name"],
                "data": p["data"],
                "col_count": struct["total_cols"],
                "struct": struct
            })
        
        return sorted(
            patterns_with_cols,
            key=lambda x: x["col_count"],
            reverse=(order == "desc")
        )

    def _group_by_column_range(self, sorted_patterns):
        groups = {"6-7": [], "8-10": []}
        for p in sorted_patterns:
            if 6 <= p["col_count"] <= 7:
                groups["6-7"].append(p)
            else:
                groups["8-10"].append(p)
        return [g for g in groups.values() if g]

    def _format_blank_column(self, ws, start_row, end_row, col_idx):
        """格式化区块外间隔列（无框线）"""
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = self.white_fill
            cell.border = self.no_border
            cell.alignment = self.center_align
            ws.row_dimensions[row].height = 24

    def _generate_single_pattern_block(self, ws, pattern, start_row, start_col):
        """生成单个图案区块（确保首行有完整框线）"""
        struct = pattern["struct"]
        title = pattern["title"]
        col_end = start_col + struct["total_cols"] - 1

        # 1. 标题行（区块第一行，强制添加完整框线）
        ws.merge_cells(start_row=start_row, start_column=start_col,
                      end_row=start_row, end_column=col_end)
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = self.bold_font
        title_cell.fill = self.title_total_fill
        title_cell.alignment = self.center_align
        title_cell.border = self.block_border  # 区块首行明确应用完整框线
        ws.row_dimensions[start_row].height = 24

        # 2. 表头行
        header_row = start_row + 1
        for col_offset in range(struct["total_cols"]):
            col = start_col + col_offset
            cell = ws.cell(row=header_row, column=col, value=struct["headers"][col_offset])
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.block_border
            ws.column_dimensions[get_column_letter(col)].width = 7.5
        ws.row_dimensions[header_row].height = 24

        # 3. 数据行
        data_start = header_row + 1
        data_end = data_start + struct["row_count"] - 1
        for row_offset, data_row in enumerate(struct["data"]):
            data_row_num = data_start + row_offset
            fill = self.header_fill if (row_offset % 2 == 1) else self.white_fill
            for col_offset in range(struct["total_cols"]):
                col = start_col + col_offset
                cell = ws.cell(row=data_row_num, column=col, value=data_row[col_offset])
                if col_offset == 0 or (col_offset + 1 == struct["total_col_idx"]):
                    cell.font = self.bold_font
                else:
                    cell.font = self.normal_font
                cell.fill = fill
                cell.alignment = self.center_align
                cell.border = self.block_border
            ws.row_dimensions[data_row_num].height = 24

        # 4. 合计行
        total_row = data_end + 1
        # 合计标题
        total_title_cell = ws.cell(row=total_row, column=start_col, value="合计")
        total_title_cell.font = self.bold_font
        total_title_cell.fill = self.title_total_fill
        total_title_cell.alignment = self.center_align
        total_title_cell.border = self.block_border
        ws.row_dimensions[total_row].height = 24

        # 规格列合计
        for col_offset in range(struct["spec_count"]):
            col = start_col + 1 + col_offset
            if col > col_end:
                break
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row, column=col,
                value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            )
            cell.font = self.bold_font
            cell.fill = self.title_total_fill
            cell.alignment = self.center_align
            cell.border = self.block_border

        # 行合计列的总合计
        total_col = start_col + struct["total_col_idx"] - 1
        if total_col <= col_end:
            total_col_letter = get_column_letter(total_col)
            cell = ws.cell(
                row=total_row, column=total_col,
                value=f"=SUM({total_col_letter}{data_start}:{total_col_letter}{data_end})"
            )
            cell.font = self.bold_font
            cell.fill = self.title_total_fill
            cell.alignment = self.center_align
            cell.border = self.block_border

        # 空白列
        for col_offset in range(struct["total_cols"]):
            col = start_col + col_offset
            if (col_offset == 0 
                or (1 <= col_offset <= struct["spec_count"]) 
                or (col_offset + 1 == struct["total_col_idx"])):
                continue
            cell = ws.cell(row=total_row, column=col)
            cell.fill = self.title_total_fill
            cell.border = self.block_border
            cell.alignment = self.center_align

        return total_row  # 返回当前区块的结束行

    def _generate_type_pattern_sheet(self, ws, type_name, type_data):
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.8, bottom=0.8)
        ws.title = f"{type_name}明细"

        pattern_groups = type_data.groupby("图案", as_index=False)
        patterns = [{"name": name, "data": group} for name, group in pattern_groups]
        sorted_patterns = self._sort_patterns_by_columns(patterns, order="asc")
        pattern_groups = self._group_by_column_range(sorted_patterns)
        
        current_row = 1
        date_str = self._get_formatted_date()

        for group in pattern_groups:
            for i in range(0, len(group), 2):
                # 左侧图案
                pattern1 = group[i]
                pattern1["title"] = f"{type_name} - {pattern1['name']} - {date_str}"
                left_end_row = self._generate_single_pattern_block(
                    ws, pattern1, current_row, self.left_start_col
                )

                # 右侧图案（若存在）
                right_end_row = left_end_row
                if (i + 1) < len(group):
                    pattern2 = group[i + 1]
                    pattern2["title"] = f"{type_name} - {pattern2['name']} - {date_str}"
                    right_end_row = self._generate_single_pattern_block(
                        ws, pattern2, current_row, self.right_start_col
                    )

                # 计算下一个区块的起始行
                current_row = max(left_end_row, right_end_row) + self.block_spacing

                # 绘制区块外间隔列（无框线）
                space_col = self.left_start_col + pattern1["struct"]["total_cols"]
                self._format_blank_column(ws, current_row - self.block_spacing, 
                                         max(left_end_row, right_end_row), space_col)
                ws.column_dimensions[get_column_letter(space_col)].width = self.space_col_width

                # 清理区块外的单元格（确保无框线）
                for row in range(current_row - self.block_spacing, current_row):
                    for col in range(self.left_start_col + pattern1["struct"]["total_cols"] + 1, 
                                    self.right_start_col):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self.no_border

    def generate(self, input_file=None):
        input_path = input_file or PathConfig.INPUT_FILE
        standardizer = DataStandardizer()
        df = standardizer.standardize(input_path)

        wb = Workbook()

        # 生成类型汇总表
        type_aggregation = self._calculate_type_aggregation(df)
        ws_type = wb.active
        self._generate_original_type_sheet(ws_type, type_aggregation)

        # 为每个类型生成图案明细工作表
        for type_name, type_group in df.groupby("类型"):
            ws = wb.create_sheet()
            self._generate_type_pattern_sheet(ws, type_name, type_group)

        output_path = PathConfig.get_output_path()
        wb.save(output_path)
        print(f"表格已生成：{output_path}")
        return output_path

if __name__ == "__main__":
    try:
        generator = CompleteTableGenerator()
        generator.generate()
    except Exception as e:
        print(f"生成失败：{str(e)}")
